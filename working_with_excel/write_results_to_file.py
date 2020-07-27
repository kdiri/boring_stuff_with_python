"""
.. module:: working_with_excel.read_from_file
   :synopsis: Get the information from excel file and write computing results to another file.
"""

from openpyxl import load_workbook
import pprint
from loguru import logger
from typing import Dict


def load_census_file():
    return load_workbook("excel_files/censuspopdata.xlsx")["Population by Census Tract"]


def compute_population_of_each_country(sheet: load_workbook):
    new_dict: dict = {}
    for state, county, pop in zip(sheet["B"], sheet["C"], sheet["D"]):
        pop = int(pop.value) if str(pop.value).isdigit() else 0
        if state.value in new_dict:
            if county.value in new_dict[state.value]:
                new_dict[state.value][county.value]["pop"] += pop
            else:
                new_dict[state.value].update({county.value: {"pop": pop}})
        else:
            new_dict[state.value] = {county.value: {"pop": pop}}
    return new_dict


def write_results_to_file(result_dict: Dict):
    logger.info(f"| ----- Results will be written into file.")
    with open("census2010.py", "w") as res_file:
        res_file.write(f"'Results': {pprint.pformat(result_dict)}")
    logger.success(f"| ----- Results are in the census2010.py file.")


def handle():
    sheet = load_census_file()
    results = compute_population_of_each_country(sheet)
    write_results_to_file(results)


if __name__ == "__main__":
    handle()
