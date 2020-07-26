"""
.. module:: working_with_excel.read_from_file
   :synopsis: Get the information from excel file.
"""

from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


def get_header_info(wb: load_workbook):
    logger.info(f"Column string from index: {get_column_letter(2)}")
    logger.info(f"Column string from index: {get_column_letter(27)}")
    logger.info(f"Column string from index: {get_column_letter(900)}")
    logger.info(f"Column index of A: {column_index_from_string('A')}")
    logger.info(f"Column index of AA: {column_index_from_string('AA')}")
    sheet = wb["Sheet1"]
    logger.info(f"Max column letter: {get_column_letter(sheet.max_column)}")
    logger.info(f"All sheets from A1 to C3: {tuple(sheet['A1':'C3'])}")

    for row_cells in sheet["A1":"C3"]:
        for row_cell in row_cells:
            logger.info(f"Cell info: {row_cell.coordinate}:{row_cell.value}")
        logger.info("| ------ End of row.")


def load_file():
    wb = load_workbook("excel_files/example.xlsx")
    get_header_info(wb)


if __name__ == "__main__":
    load_file()
